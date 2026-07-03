#!/usr/bin/env python3
"""
Apply Season Manager bulk actions to the Law & Order tracker workbook.

Use this when you prefer doing bulk marking inside Excel:
1) Open Law_Order_Professional_Watch_Tracker.xlsx
2) Go to Season Manager
3) Set Bulk Action for one or more rows
4) Save and close the workbook
5) Run: python apply_bulk_actions_to_excel.py

Requirements:
    pip install openpyxl
"""

from pathlib import Path
from openpyxl import load_workbook

WORKBOOK_PATH = Path("Law_Order_Professional_Watch_Tracker.xlsx")
OUTPUT_PATH = Path("Law_Order_Professional_Watch_Tracker_Bulk_Updated.xlsx")

VALID_ACTIONS = {
    "Mark Watched": "Watched",
    "Mark Unwatched": "Not Started",
    "Mark Skipped": "Skipped",
}

def main():
    if not WORKBOOK_PATH.exists():
        raise SystemExit(f"Workbook not found: {WORKBOOK_PATH}")

    wb = load_workbook(WORKBOOK_PATH)
    guide = wb["Chronological Guide"]
    manager = wb["Season Manager"]

    gh = [c.value for c in guide[1]]
    mh = [c.value for c in manager[1]]
    status_col = gh.index("Status") + 1
    show_col = gh.index("Show") + 1
    season_col = gh.index("Season") + 1
    m_show_col = mh.index("Show") + 1
    m_season_col = mh.index("Season") + 1
    m_action_col = mh.index("Bulk Action") + 1

    actions = []
    for row in range(2, manager.max_row + 1):
        action = manager.cell(row=row, column=m_action_col).value
        if action in VALID_ACTIONS:
            actions.append((manager.cell(row=row, column=m_show_col).value, manager.cell(row=row, column=m_season_col).value, VALID_ACTIONS[action], row))

    changed = 0
    for show, season, new_status, mrow in actions:
        for row in range(2, guide.max_row + 1):
            if guide.cell(row=row, column=show_col).value == show and str(guide.cell(row=row, column=season_col).value) == str(season):
                guide.cell(row=row, column=status_col).value = new_status
                changed += 1
        manager.cell(row=mrow, column=m_action_col).value = "No Action"

    wb.save(OUTPUT_PATH)
    print(f"Applied {len(actions)} season actions and updated {changed} episodes.")
    print(f"Saved: {OUTPUT_PATH}")

if __name__ == "__main__":
    main()
