#!/usr/bin/env python3
"""Fetch current Trakt episode catalogs and append any missing episodes to the web app data + Excel tracker.
Requires trakt_config.json. It does not change watched status; run sync_trakt_and_excel.py afterward.
"""
from __future__ import annotations
import json, re, time
from pathlib import Path
from datetime import datetime
import requests
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from sync_trakt_and_excel import BASE_URL, SHOW_SLUGS, headers, load_config

EPISODES_JS = Path("law_order_tracker_app/data/episodes.js")
WORKBOOK_PATH = Path("Law_Order_Professional_Watch_Tracker.xlsx")
OUTPUT_WORKBOOK = Path("Law_Order_Professional_Watch_Tracker_Updated_Catalog.xlsx")

def get_json(cfg, path):
    r = requests.get(f"{BASE_URL}{path}", headers=headers(cfg), timeout=60)
    r.raise_for_status()
    return r.json()

def load_app_episodes():
    txt = EPISODES_JS.read_text(encoding="utf-8")
    return json.loads(txt[txt.find("["):txt.rfind("]")+1])

def save_app_episodes(eps):
    eps.sort(key=lambda e: (e.get("airDate") or "9999-99-99", str(e.get("show")), int(e.get("season") or 0), int(e.get("episode") or 0)))
    for i, e in enumerate(eps, 1):
        e["order"] = i
        e["id"] = f"{e['show']}|{e['season']}|{e['episode']}|{i}"
    EPISODES_JS.write_text("window.LAW_ORDER_EPISODES = " + json.dumps(eps, indent=2, ensure_ascii=False) + ";\n", encoding="utf-8")

def fetch_show_episodes(cfg, show, slug):
    seasons = get_json(cfg, f"/shows/{slug}/seasons?extended=episodes,full")
    out=[]
    for season in seasons:
        sn = season.get("number")
        if not isinstance(sn, int) or sn == 0:
            continue
        for ep in season.get("episodes", []) or []:
            en = ep.get("number")
            if not isinstance(en, int):
                continue
            first = ep.get("first_aired") or ""
            air = first[:10] if len(first)>=10 else ""
            # Skip unaired/future placeholder rows if Trakt has no date yet.
            if not air:
                continue
            out.append({
                "id":"", "order":0, "status":"Not Started", "airDate":air, "show":show,
                "season":sn, "episode":en, "code":f"{sn:02d}.{en:02d}",
                "title":ep.get("title") or "", "notes":"Added from Trakt catalog refresh",
                "sourceWatch":"", "era":"Trakt catalog refresh", "sourceTab":"Trakt API"
            })
    return out

def update_excel(eps):
    wb = load_workbook(WORKBOOK_PATH)
    ws = wb["Chronological Guide"]
    headers = [c.value for c in ws[1]]
    # Clear and rewrite rows to avoid order/id mismatches.
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row-1)
    for e in eps:
        row=[e.get('order'), e.get('status','Not Started'), e.get('airDate'), e.get('show'), e.get('season'), e.get('episode'), e.get('code'), e.get('title'), e.get('notes'), e.get('sourceWatch'), e.get('era'), e.get('sourceTab')]
        ws.append(row)
    colors = {
        "Law & Order":"b91c1c","Law & Order: Special Victims Unit":"2563eb","Law & Order: SVU":"2563eb","Law & Order: Criminal Intent":"7c3aed","Homicide: Life on the Street":"d97706","Law & Order: Organized Crime":"0e7490","Criminal Intent: Toronto":"9333ea","Law & Order: UK":"dc2626","Law & Order: LA":"f97316","Law & Order True Crime":"525252","Law & Order: Trial by Jury":"059669","Conviction":"db2777","Deadline":"ca8a04","New York Undercover":"0891b2"}
    for r in range(2, ws.max_row+1):
        show=ws.cell(r,4).value
        ws.cell(r,4).fill=PatternFill('solid', fgColor=colors.get(show,'64748b'))
    wb.save(OUTPUT_WORKBOOK)

def main():
    cfg=load_config()
    eps=load_app_episodes()
    existing={(e['show'], int(e['season']), int(e['episode'])) for e in eps if e.get('season') and e.get('episode')}
    added=[]
    for show, slug in SHOW_SLUGS.items():
        print(f"Scanning Trakt catalog for {show}...")
        try:
            catalog=fetch_show_episodes(cfg, show, slug)
        except Exception as exc:
            print(f"  WARNING: failed: {exc}"); continue
        for e in catalog:
            key=(e['show'], int(e['season']), int(e['episode']))
            if key not in existing:
                eps.append(e); existing.add(key); added.append(e)
    save_app_episodes(eps)
    update_excel(eps)
    Path("catalog_refresh_report.json").write_text(json.dumps({"added_count":len(added),"added":added},indent=2,ensure_ascii=False),encoding="utf-8")
    print(f"Done. Added {len(added)} missing episodes. Updated {EPISODES_JS} and {OUTPUT_WORKBOOK}.")
if __name__ == "__main__": main()
