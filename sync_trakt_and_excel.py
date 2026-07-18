#!/usr/bin/env python3
"""Generate the shared watched_status.json from the complete episodes.js guide.

This legacy/shared workflow is optional. Personal users should normally sign in
with Trakt and sync through Supabase, which never commits or redeploys the site.

The script now scans the full Wolf Universe catalog instead of the old 1,796-row
Excel workbook. Excel output can still be requested explicitly with
--update-excel for archival use.
"""
from __future__ import annotations

import argparse
import json
import os
import re
import time
from pathlib import Path
from typing import Any

import requests

ROOT = Path(__file__).resolve().parent
BASE_URL = "https://api.trakt.tv"
CONFIG_PATH = ROOT / "trakt_config.json"
TOKEN_PATH = ROOT / "trakt_token.json"
EPISODES_JS = ROOT / "law_order_tracker_app" / "data" / "episodes.js"
APP_STATUS_PATH = ROOT / "law_order_tracker_app" / "data" / "watched_status.json"
ROOT_STATUS_PATH = ROOT / "law_order_watch_status_from_trakt.json"
DEBUG_PATH = ROOT / "law_order_tracker_app" / "data" / "trakt_sync_debug.json"
USER_AGENT = "Wolf-Universe-Watch-Tracker/4.0"


def norm_text(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").lower().replace("&", " and ")).strip()


def norm_num(value: Any) -> str:
    try:
        return str(int(value))
    except (TypeError, ValueError):
        return str(value or "").strip()


ALIASES = {
    norm_text("Law & Order: SVU"): norm_text("Law & Order: Special Victims Unit"),
    norm_text("SVU"): norm_text("Law & Order: Special Victims Unit"),
    norm_text("Criminal Intent"): norm_text("Law & Order: Criminal Intent"),
    norm_text("Organized Crime"): norm_text("Law & Order: Organized Crime"),
    norm_text("Chicago PD"): norm_text("Chicago P.D."),
    norm_text("NY Undercover"): norm_text("New York Undercover"),
}


def norm_show(value: Any) -> str:
    value = norm_text(value)
    return ALIASES.get(value, value)


def episode_key(show: Any, season: Any, episode: Any) -> str:
    return f"{norm_show(show)}|{norm_num(season)}|{norm_num(episode)}"


def load_window_array(path: Path, variable: str) -> list[dict]:
    text = path.read_text(encoding="utf-8")
    match = re.search(rf"window\.{re.escape(variable)}\s*=\s*(\[.*\])\s*;?\s*$", text, re.S)
    if not match:
        raise SystemExit(f"Could not parse {variable} from {path}")
    return json.loads(match.group(1))


def load_config() -> dict:
    if not CONFIG_PATH.exists():
        raise SystemExit("Missing trakt_config.json. Copy trakt_config.example.json and add your app credentials.")
    config = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    if not config.get("client_id") or not config.get("client_secret"):
        raise SystemExit("trakt_config.json must contain client_id and client_secret")
    return config


def headers(config: dict, access_token: str = "") -> dict:
    result = {
        "Content-Type": "application/json",
        "User-Agent": USER_AGENT,
        "trakt-api-version": "2",
        "trakt-api-key": config["client_id"],
    }
    if access_token:
        result["Authorization"] = f"Bearer {access_token}"
    return result


def save_token(token: dict) -> None:
    TOKEN_PATH.write_text(json.dumps(token, indent=2), encoding="utf-8")


def refresh_token(config: dict, token: dict) -> dict:
    redirect_uri = config.get("redirect_uri") or os.environ.get("TRAKT_REDIRECT_URI") or "urn:ietf:wg:oauth:2.0:oob"
    response = requests.post(
        f"{BASE_URL}/oauth/token",
        json={
            "refresh_token": token["refresh_token"],
            "client_id": config["client_id"],
            "client_secret": config["client_secret"],
            "redirect_uri": redirect_uri,
            "grant_type": "refresh_token",
        },
        headers=headers(config),
        timeout=45,
    )
    response.raise_for_status()
    refreshed = response.json()
    refreshed["created_at_local"] = int(time.time())
    save_token(refreshed)
    return refreshed


def device_auth(config: dict) -> dict:
    response = requests.post(
        f"{BASE_URL}/oauth/device/code",
        json={"client_id": config["client_id"]},
        headers=headers(config),
        timeout=45,
    )
    response.raise_for_status()
    device = response.json()
    print(f"Open {device['verification_url']} and enter code {device['user_code']}")
    interval = int(device.get("interval", 5))
    deadline = time.time() + int(device.get("expires_in", 600))
    while time.time() < deadline:
        time.sleep(interval)
        poll = requests.post(
            f"{BASE_URL}/oauth/device/token",
            json={
                "code": device["device_code"],
                "client_id": config["client_id"],
                "client_secret": config["client_secret"],
            },
            headers=headers(config),
            timeout=45,
        )
        if poll.status_code == 200:
            token = poll.json()
            token["created_at_local"] = int(time.time())
            save_token(token)
            return token
        if poll.status_code in (400, 404):
            continue
        if poll.status_code == 418:
            raise SystemExit("Trakt authorization was denied")
        poll.raise_for_status()
    raise SystemExit("Timed out waiting for Trakt authorization")


def get_token(config: dict) -> dict:
    if not TOKEN_PATH.exists():
        return device_auth(config)
    token = json.loads(TOKEN_PATH.read_text(encoding="utf-8"))
    created = int(token.get("created_at", token.get("created_at_local", 0)) or 0)
    expires = int(token.get("expires_in", 0) or 0)
    if created and expires and time.time() > created + expires - 600:
        return refresh_token(config, token)
    return token


def trakt_get(config: dict, token: dict, path: str) -> requests.Response:
    for attempt in range(3):
        response = requests.get(f"{BASE_URL}{path}", headers=headers(config, token["access_token"]), timeout=90)
        if response.status_code == 401 and attempt == 0:
            token.update(refresh_token(config, token))
            continue
        if response.status_code in (429, 500, 502, 503, 504) and attempt < 2:
            time.sleep(int(response.headers.get("Retry-After", "0") or 0) or (attempt + 1) * 2)
            continue
        response.raise_for_status()
        return response
    raise RuntimeError(f"Trakt request failed: {path}")


def id_values(ids: dict | None) -> list[str]:
    ids = ids or {}
    return [f"{key}:{ids[key]}" for key in ("trakt", "tmdb", "tvdb", "imdb", "slug") if ids.get(key) not in (None, "")]


def build_indexes(guide: list[dict]) -> dict[str, dict[str, list[dict]]]:
    by_name: dict[str, list[dict]] = {}
    by_show_id: dict[str, list[dict]] = {}
    by_episode_id: dict[str, list[dict]] = {}
    for ep in guide:
        season, number = norm_num(ep.get("season")), norm_num(ep.get("episode"))
        for show_name in {norm_show(ep.get("show")), norm_show(ep.get("titleShow"))} - {""}:
            by_name.setdefault(f"{show_name}|{season}|{number}", []).append(ep)
        for show_id in id_values(ep.get("showTraktIds")) + ([f"slug:{ep['traktSlug']}"] if ep.get("traktSlug") else []):
            by_show_id.setdefault(f"{show_id}|{season}|{number}", []).append(ep)
        for episode_id in id_values(ep.get("traktIds")):
            by_episode_id.setdefault(episode_id, []).append(ep)
    return {"name": by_name, "show_id": by_show_id, "episode_id": by_episode_id}


def match_episode(indexes: dict, show: dict, episode: dict) -> list[dict]:
    matches: dict[str, dict] = {}
    season, number = episode.get("season"), episode.get("number")
    for episode_id in id_values(episode.get("ids")):
        for ep in indexes["episode_id"].get(episode_id, []):
            matches[str(ep["id"])] = ep
    for show_id in id_values(show.get("ids")):
        for ep in indexes["show_id"].get(f"{show_id}|{norm_num(season)}|{norm_num(number)}", []):
            matches[str(ep["id"])] = ep
    key = episode_key(show.get("title"), season, number)
    for ep in indexes["name"].get(key, []):
        matches[str(ep["id"])] = ep
    return list(matches.values())


def fetch_watched(config: dict, token: dict, guide: list[dict]) -> tuple[dict, dict]:
    indexes = build_indexes(guide)
    matched: dict[str, dict] = {}
    debug: dict[str, Any] = {"sources": {}, "unmatched": [], "generatedAt": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())}

    watched_shows = trakt_get(config, token, "/sync/watched/shows?extended=full").json()
    debug["sources"]["watched_shows"] = len(watched_shows)
    for item in watched_shows:
        show = item.get("show") or {}
        for season in item.get("seasons") or []:
            for raw_episode in season.get("episodes") or []:
                episode = {**raw_episode, "season": season.get("number")}
                found = match_episode(indexes, show, episode)
                if not found:
                    debug["unmatched"].append({"show": show.get("title"), "season": episode.get("season"), "episode": episode.get("number")})
                for ep in found:
                    matched[str(ep["id"])] = ep

    try:
        history = trakt_get(config, token, "/sync/history/episodes?limit=10000&extended=full").json()
        debug["sources"]["history_episodes"] = len(history)
        for item in history:
            show, episode = item.get("show") or {}, item.get("episode") or {}
            for ep in match_episode(indexes, show, episode):
                matched[str(ep["id"])] = ep
    except Exception as exc:
        debug["sources"]["history_error"] = str(exc)

    return matched, debug


def write_outputs(guide: list[dict], matched: dict[str, dict], debug: dict) -> None:
    statuses: dict[str, str] = {}
    episodes: list[dict] = []
    by_show: dict[str, int] = {}
    for ep in guide:
        if str(ep.get("id")) not in matched:
            continue
        statuses[str(ep["id"])] = "Watched"
        statuses[episode_key(ep.get("show"), ep.get("season"), ep.get("episode"))] = "Watched"
        episodes.append({
            "id": str(ep["id"]), "show": ep.get("show"), "season": ep.get("season"),
            "episode": ep.get("episode"), "status": "Watched"
        })
        by_show[ep.get("show") or "Unknown"] = by_show.get(ep.get("show") or "Unknown", 0) + 1

    payload = {
        "version": 8,
        "source": "trakt-shared-full-guide",
        "exportedAt": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
        "summary": {
            "guideRowsScanned": len(guide),
            "guideRowsMarkedWatched": len(episodes),
            "watchedCountsByShow": dict(sorted(by_show.items())),
        },
        "statuses": statuses,
        "episodes": episodes,
    }
    APP_STATUS_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    ROOT_STATUS_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    debug["summary"] = payload["summary"]
    debug["unmatched"] = debug.get("unmatched", [])[:500]
    DEBUG_PATH.write_text(json.dumps(debug, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Matched {len(episodes)}/{len(guide)} guide entries. Wrote {APP_STATUS_PATH}")


def update_excel_legacy(matched: dict[str, dict]) -> None:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit("Install openpyxl to use --update-excel")
    source = ROOT / "Law_Order_Professional_Watch_Tracker.xlsx"
    output = ROOT / "Law_Order_Professional_Watch_Tracker_Trakt_Synced.xlsx"
    if not source.exists():
        print("Legacy workbook not found; skipping Excel output")
        return
    workbook = load_workbook(source)
    if "Chronological Guide" not in workbook.sheetnames:
        print("Legacy workbook has no Chronological Guide sheet; skipping")
        return
    sheet = workbook["Chronological Guide"]
    headers = {str(cell.value): index + 1 for index, cell in enumerate(sheet[1])}
    needed = {"Status", "Show", "Season", "Episode"}
    if not needed.issubset(headers):
        print("Legacy workbook headers are incompatible; skipping")
        return
    guide_keys = {(norm_show(ep.get("show")), norm_num(ep.get("season")), norm_num(ep.get("episode"))) for ep in matched.values()}
    for row in range(2, sheet.max_row + 1):
        key = (
            norm_show(sheet.cell(row, headers["Show"]).value),
            norm_num(sheet.cell(row, headers["Season"]).value),
            norm_num(sheet.cell(row, headers["Episode"]).value),
        )
        sheet.cell(row, headers["Status"]).value = "Watched" if key in guide_keys else "Not Started"
    workbook.save(output)
    print(f"Wrote optional legacy workbook: {output}")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--update-excel", action="store_true", help="also update the legacy workbook")
    args = parser.parse_args()
    config = load_config()
    token = get_token(config)
    guide = load_window_array(EPISODES_JS, "LAW_ORDER_EPISODES")
    matched, debug = fetch_watched(config, token, guide)
    write_outputs(guide, matched, debug)
    if args.update_excel:
        update_excel_legacy(matched)


if __name__ == "__main__":
    main()
